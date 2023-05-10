# |--------------------------------------------------------------------------------------------|
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |                                                                                            |
# | Module Name: app.py - OFFLINE ONLY                                                         |
# | Author: Noah Shields                                                                       |
# | Date: 05/15/2023                                                                           |
# |                                                                                            |
# | Description: This is the driving program for this application. It includes all needed      |
# |              classes, flask routes, helper methods, etc.                                   |
# |              All methods will be described below, or in the README file                    |
# |              This is the OFFLINE version only, a guide to make it online in the the README |
# |                                                                                            |
# | Dependencies: This uses a flask backend and utitizes bootstrap and jinja to display and    |
# |               allow the user to guide through the flask routes                             |
# |                                                                                            |
# | Notes: If making modifications to this code read the README file as it explains each       |
# |        data structure and algorithm in detail. The comments below are supplemental only    |
# |                                                                                            |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *~|
# |--------------------------------------------------------------------------------------------|

from flask import Flask, render_template, request, redirect, url_for
from flask_bootstrap import Bootstrap
from flask_moment import Moment
from flask_wtf import FlaskForm
from wtforms import widgets, StringField, SubmitField, SelectField, Form, SelectMultipleField
from wtforms.validators import DataRequired
from flask_login import LoginManager, UserMixin, login_required, login_user, logout_user
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import ast, docx, datetime, os, json, random
#from db import connection


app = Flask(__name__)

#this is the login manager. All routes decorated with @login_required,
#are protected by this manager. A logout route is also optional, but not implemented
login_manager = LoginManager()
login_manager.init_app(app)

app.config['SECRET_KEY'] = 'super_secret'
bootstrap = Bootstrap(app)
moment = Moment(app)
#This connection is used for the connection to a outside database
#conn = connection(host = '', database = '', user = '', password = '', port = '')

#dictionary of tutors. The key is the name of the tutor, and the value is a Tutor object,
#see the Tutor object below
tutors = {}

#dictionary to be used only for the word and excel sheets
#it is similar to all_courses, but it includes the names as an additional layer. here is an example
parsedData = {}

#set of course codes
#used as an organized way of parsed data, used whenever one needs to know something about a course, but not the tutors
all_courses = {}

#A tutor has a name, a set of courses they can tutor, and a set of times they work
class Tutor:
    def __init__(self, name=None, courses=None, availabilities=None):
        self.name = name
        self.courses = set(courses) if courses else set()
        self.availabilities = set(availabilities) if availabilities else set()

# |--------------------------------------------------------------------------------------------|
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Classes ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |--------------------------------------------------------------------------------------------|
# |These are flask form classes, they are used to create the WTF forms that you see            |
# |Sepcifically the interactive portions... The submit buttons, and the select/string fields   |
# |--------------------------------------------------------------------------------------------|

# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Home and Login Classes ~ * ~ * ~ * ~ * ~ * ~ * ~
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
class User(UserMixin):
    def __init__(self, id):
        self.id = id

class LoginForm(FlaskForm):
    user = StringField('VID:', validators=[DataRequired()])
    password = StringField('PASSWORD:', validators=[DataRequired()])
    submit = SubmitField('Submit')
    cond = False
    
class AdminSelectionForm(FlaskForm):
    insert = SubmitField('Insert')
    delete = SubmitField('Delete')
    view = SubmitField('view')
    update = SubmitField('update')

# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Delete Classes ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
class DeleteTutorForm(FlaskForm):
    name = SelectField('Name', choices = [], validators=[DataRequired()])
    submit = SubmitField('Submit')

class DeleteCourseForm(FlaskForm):
    course_code = SelectField('Course Code', choices = [], validators=[DataRequired()])
    submit = SubmitField('Submit')

class DeleteUserForm(FlaskForm):
    user = SelectField('Username', validators=[DataRequired()])
    submit = SubmitField('Submit')

#class DeleteExpertiseForm(FlaskForm)       ##Same form as insert, so we reuse it
#class DeleteAvailabilityForm(FlaskForm)    ##Same form as insert, so we reuse it

# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Insert Classes ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
class TutorForm(FlaskForm):
    name = StringField('Name', validators=[DataRequired()])
    submit = SubmitField('Submit')

class CourseForm(FlaskForm):
    course_code = StringField('Course Code', validators=[DataRequired()])
    submit = SubmitField('Submit')

class ExpertiseForm(FlaskForm):
    name = SelectField('Name', choices = [], validators=[DataRequired()])
    course_codes = SelectMultipleField('course code', choices = [], validators=[DataRequired()])
    submit = SubmitField('Submit')

class AvailabilityForm(FlaskForm):
    name = SelectField('Name', choices = [], validators=[DataRequired()])
    days = SelectMultipleField('Days', choices = [], validators=[DataRequired()])
    times = SelectMultipleField('Times', choices = [], validators=[DataRequired()])
    submit = SubmitField('Submit')

class OldDataForm(FlaskForm):
    oldData = StringField('Existing Data', validators=[DataRequired()])
    submit = SubmitField('Submit')

class UserForm(FlaskForm):
    user = StringField('Username', validators=[DataRequired()])
    password = StringField('password', validators=[DataRequired()])
    submit = SubmitField('Submit')



# |--------------------------------------------------------------------------------------------|
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Routing ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |--------------------------------------------------------------------------------------------|
# |All the main flask routes, that are matched with some HTML page                             |
# |--------------------------------------------------------------------------------------------|

# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Error Handling ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_server_error(e):
    return render_template('500.html'), 500


# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Login Manager ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# Load user
@login_manager.user_loader
def load_user(user_id):
    users = getUsers()
    if user_id in users:
        return User(user_id)
    return None

# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Login and Home Pages ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~
@app.route('/', methods=['GET', 'POST'])
def home():
    return render_template("home.html")

# Login home
@app.route('/login', methods=['GET', 'POST'])
def login():
    vid = None
    password = None
    form = LoginForm()
    #cond = True
    if form.validate_on_submit():
        user = form.user.data
        password = form.password.data
        users = getUsers()

        #check if the user exists and if the password matches, if so, log them in
        if user in users and users[user] == password:
            user = User(user)
            login_user(user)

            insertOldData()  #Extract all data from oldData.txt, and populate the data structutes
                             #This allows for the user to pick up where they left off

            return redirect(url_for('selections'))
    return render_template('login.html', form = form)

#simply the selections of tools... insert, delete, view, update, create tables, reset, look for previous data
@app.route('/selections', methods=['GET', 'POST'])
@login_required
def selections():
    return render_template("admin.html")

# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Delete Pages ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
@app.route('/delete', methods=['GET', 'POST'])
@login_required
def delete():
    return render_template('delete.html')

@app.route('/delete_tutor', methods=['GET', 'POST'])
@login_required
def delete_tutor():
    name = None
    form = DeleteTutorForm()

    names = sorted([(tutor[0], tutor[1].name) for tutor in tutors.items()]) #sorted choices of all names

    form.name.choices = names

    if form.validate_on_submit():
        name = form.name.data
        if name in tutors.keys():   #if the tutor exists
            tutors.pop(name)        #delete tutor from tutors list
            update_all_courses()    #update all_courses
            updateParsedData()      #updated parsed data
            return render_template('delete_tutor.html', form = form)

    return render_template('delete_tutor.html', form = form)

@app.route('/delete_course', methods=['GET', 'POST'])
@login_required
def delete_course():
    course_code = None
    all_course_codes = None
    tempCodes = set()
    form = DeleteCourseForm()

    for subject in all_courses.keys():
        for number in all_courses[subject]:
            tempCodes.add(f"{subject}{number}")

    all_course_codes = sorted([(code,code) for code in tempCodes])    #sorted choices of all courses

    form.course_codes.choices = all_course_codes

    if form.validate_on_submit():
        course = form.course.data
        subject = course[:4]            #CMSC in CMSC101
        number = course[4:]             #101 in CMSC101
        if subject not in parsedData or number not in parsedData[subject]:  #if this course doesnt exist
            return render_template('delete_course.html', form = form)

        tutors = [data[0] for data in parsedData[subject][number]] #grabs a list of all tutors that tutors a subject
        for tutor in tutors:
            tutors[tutor].courses.discard(course) #delete course from tutors list of courses
        
        if subject in all_courses:
            all_courses[subject].pop(number)    #delete course from all_courses, which updates all_courses

        updateParsedData()  #update parsed data        

    return render_template('delete_course.html', form = form)

#Really should only be used with a database
@app.route('/delete_user', methods=['GET', 'POST'])
@login_required
def delete_user():
    # user = None
    # password = None
    # form = UserForm()

    # if form.validate_on_submit():
    #     user = form.user.data
    #     password = form.password.data
    #     conn.delete({"table":"user","data":{"user":user,"password":password}})

    return render_template('delete_user.html')


# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Update Pages ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
#options for updating, expertise and availabilities
@app.route('/update', methods=['GET', 'POST'])
@login_required
def update():
    return render_template('update.html')

#this is basically a copy of insert_expertise, but the expertise of the tutor is first reset
@app.route('/update_expertise', methods=['GET', 'POST'])
@login_required
def update_expertise():
    global tutors
    names = None
    course_codes = None
    all_course_codes = None
    tempCodes = set()
    form = ExpertiseForm()

    for subject in all_courses.keys():
        for number in all_courses[subject]:
            tempCodes.add(f"{subject}{number}")

    names = sorted([(tutor[0], tutor[1].name) for tutor in tutors.items()]) #choices of all names
    all_course_codes = sorted([(code,code) for code in tempCodes])         #choices of all courses
    
    form.course_codes.choices = all_course_codes
    form.name.choices = names


    if form.validate_on_submit():
        course_codes = form.course_codes.data
        name = form.name.data
        tutors[name].courses = set()    #expertise of the tutor are reset
        for course in course_codes:
            tutors[name].courses.add(course)

        update_all_courses()
        updateParsedData()
            
            


        
    
    return render_template('update_expertise.html', form = form)

#this is basically a copy of insert_availabilties, but the availabilties of the tutor is first reset
@app.route('/update_availability', methods=['GET', 'POST'])
@login_required
def update_availability():
    global tutors
    times = None
    days = None
    form = AvailabilityForm()

    names = sorted([(tutor[0], tutor[1].name) for tutor in tutors.items()])

    form.name.choices = names
    form.days.choices = [('sunday', 'Sunday'), ('monday', 'Monday'), ('tuesday', 'Tuesday'), ('wednesday', 'Wednesday'), ('thursday', 'Thursday'), ('friday', 'Friday'), ('saturday', 'Saturday')]
    form.times.choices = [('9', '9:00 - 10:00 AM'), ('10', '10:00 - 11:00 AM'), ('11', '11:00 - 12:00 PM'), ('12', '12:00 - 1:00 PM'), ('13', '1:00 - 2:00 PM'), ('14', '2:00 - 3:00 PM'), ('15', '3:00 - 4:00 PM'), ('16', '4:00 - 5:00 PM'), ('17', '5:00 - 6:00 PM'), ('18', '6:00 - 7:00 PM'), ('19', '7:00 - 8:00 PM'), ('20', '8:00 - 9:00 PM'), ('21', '9:00 - 10:00 PM')]


    if form.validate_on_submit():
        name = form.name.data
        days = form.days.data
        times = form.times.data
        tutors[name].availabilities = set() #eavailabilities are reset
        for day in days:
            for time in times:
                tutors[name].availabilities.add((day,time))

        update_all_courses()
        updateParsedData()
    
    return render_template('update_availability.html', form = form)


# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Select Pages ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
@app.route('/select', methods=['GET', 'POST'])
@login_required
def select():
    return render_template('select.html')

@app.route('/select_by_tutor', methods=['GET', 'POST'])
@login_required
def select_by_tutor():
    data = []
    headers = ('Name', 'Courses', 'Availability')
    for tutor in tutors.values():
        name = tutor.name
        courses = ', '.join(tutor.courses)
        

        #get the tutors availabilites and sort them
        availabilities = sort_availabilities(tutor.availabilities)

        #make them readable... put them into ranges based on Day... monday 2pm-3pm
        availabilities = combineDays(availabilities)

        #add to table
        data.append((name, courses, availabilities))        

    #the select_table page is used for any type of table, this is used whenever a table is to be shown
    return render_template('select_table.html', headers = headers, data = data)

@app.route('/select_by_course', methods=['GET', 'POST'])
@login_required
def select_by_course():
    headers = ('Course', 'Tutors', 'Times')
    data = []
    updateParsedData() #update parsed data to the most recent

    #parsedData is better to use here as it is parsed in a way where everything is organixzed by course rather than tutor
    for subject in parsedData.items():
        course_sub = subject[0]                     #course subject CMSC in CMSC101
        for course in subject[1].items():           #for each course subject...
            tutors = ""                             #list of tutors that teach that course
            availabilities = set({})                #list of times a course is taught
            course_num = course[0]                  #course number 101 in CMSC101
            course_name = course_sub + course_num   #CMSC101
            for course_data in course[1]:           #for each tutor that teaches that course
                tutors += f"{course_data[0]}, "     #add them to list of tutors
                availabilities = availabilities | course_data[1] #add their availabilities
                
            availabilities = sort_availabilities(availabilities)    #sort the availabilites
            availabilities = combineDays(availabilities)            #make it readable in a range... monday 2pm-3pm
            data.append((course_name, tutors, availabilities))      #add to table


    return render_template('select_table.html', headers = headers, data = data)


# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Insert Pages ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ *
# ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ 
#shows options of things to insert... tutor, expertise, course, availabilties, user, old data
@app.route('/insert', methods=['GET', 'POST'])
@login_required
def insert():
    return render_template('insert.html')

@app.route('/insert_tutor', methods=['GET', 'POST'])
@login_required
def insert_tutor():
    name = None
    form = TutorForm()

    if form.validate_on_submit():
        name = form.name.data.lower()       #grab name and lowercase it
        if name in tutors.keys():           #if tutor is already inserted, return with dup == 1, indicating the user that the tutor already exists
            return render_template('insert_tutor.html', form = form, dup = 1)

        tutor = Tutor(name = name)  #otherwise make a new tutor
        tutors[name] = tutor
    return render_template('insert_tutor.html', form = form, dup = 0)

@app.route('/insert_course', methods=['GET', 'POST'])
@login_required
def insert_course():
    course_code = None
    form = CourseForm()

    if form.validate_on_submit():
        course_code = form.course_code.data.upper() #grab course name and uppercase it
        courseSubject = course_code[:4]     #split into the subject and number
        courseNumber = course_code[4:]
        if courseSubject not in all_courses:    #if its a new subject
            all_courses[courseSubject] = {}

        if courseNumber not in all_courses[courseSubject]: #if a new course in that subject
            all_courses[courseSubject][courseNumber] = set({})

    return render_template('insert_course.html', form = form)

@app.route('/insert_expertise', methods=['GET', 'POST'])
@login_required
def insert_expertise():

    names = None
    course_codes = None
    all_course_codes = None
    tempCodes = set({})
    form = ExpertiseForm()

    #grab all the courses
    for subject in all_courses.keys():
        for number in all_courses[subject]:
            tempCodes.add(f"{subject}{number}")

    #add choices for courses and names
    names = sorted([(tutor[0], tutor[1].name) for tutor in tutors.items()])
    all_course_codes = sorted([(code,code) for code in tempCodes])
    
    form.course_codes.choices = all_course_codes
    form.name.choices = names


    if form.validate_on_submit():
        course_codes = form.course_codes.data
        name = form.name.data
        for course in course_codes:
            tutors[name].courses.add(course)    #add list of courses to the tutor

        #function used to update the all_courses data structure
        update_all_courses(name)
            
            


        
    
    return render_template('insert_expertise.html', form = form)

@app.route('/insert_availability', methods=['GET', 'POST'])
@login_required
def insert_availability():
    times = None
    days = None
    form = AvailabilityForm()

    names = sorted([(tutor[0], tutor[1].name) for tutor in tutors.items()]) #list of all names

    form.name.choices = names
    form.days.choices = [('sunday', 'Sunday'), ('monday', 'Monday'), ('tuesday', 'Tuesday'), ('wednesday', 'Wednesday'), ('thursday', 'Thursday'), ('friday', 'Friday'), ('saturday', 'Saturday')]
    form.times.choices = [('9', '9:00 - 10:00 AM'), ('10', '10:00 - 11:00 AM'), ('11', '11:00 - 12:00 PM'), ('12', '12:00 - 1:00 PM'), ('13', '1:00 - 2:00 PM'), ('14', '2:00 - 3:00 PM'), ('15', '3:00 - 4:00 PM'), ('16', '4:00 - 5:00 PM'), ('17', '5:00 - 6:00 PM'), ('18', '6:00 - 7:00 PM'), ('19', '7:00 - 8:00 PM'), ('20', '8:00 - 9:00 PM'), ('21', '9:00 - 10:00 PM')]


    if form.validate_on_submit():
        name = form.name.data
        days = form.days.data
        times = form.times.data

        #add all day and time pairs into the tutors availabilities (monday, 9) is monday from 9am to 10am
        for day in days:
            for time in times:
                tutors[name].availabilities.add((day,time)) 

        #update all_courses data structure
        update_all_courses(name)
    
    return render_template('insert_availability.html', form = form)

#should only be used if database is used
@app.route('/insert_user', methods=['GET', 'POST'])
@login_required
def insert_user():
    # user = None
    # password = None
    # form = UserForm()

    # if form.validate_on_submit():
    #     user = form.user.data
    #     password = form.password.data
    #     conn.insert({"table":"user","data":{"user":user,"password":password}})

    return render_template('insert_user.html', form = form)

#There is a txt file in /deliverables that is similar to a CSV file. insertOldData takes in a string and parses it
#This parsed data is then used to update all 3 data structures
@app.route('/insert_oldData', methods=['GET', 'POST'])
@login_required
def insert_oldData():
    form = OldDataForm()

    if form.validate_on_submit():
        oldData = form.oldData.data
        insertOldData(oldData=oldData)

    return render_template('insert_oldData.html', form = form)


# |--------------------------------------------------------------------------------------------|
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ Helper Methods and Routes * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * ~ * |
# |--------------------------------------------------------------------------------------------|

@app.route('/reset')
@login_required
def reset():
    global tutors
    global all_courses
    global parsedData

    #backup the data, store in deletedData.txt, and reset all three data structures

    downloadExistingData("deletedData.txt")
    tutors = {}
    parsedData = {}
    all_courses = {} 

    try:
        os.remove("deliverables/oldData.txt")
    except FileNotFoundError:
        pass

    return redirect('/selections')

def getUsers():
    users = {'admin':'123'}
    # users = conn.select({"table":"user","data":{},"columns":["*"]})
    # users = {user[0]:user[1] for user in users}
    return users

@app.route('/createTable')
@login_required
def createTable():
    #update the parsedData data structure, download the oldData.txt file, downlaod the word and excel docs

    updateParsedData()
    downloadExistingData()
    downloadDoc()
    downloadExcel()
    return redirect('/selections')

@app.route('/automatic_insert_old_data')
def automatic_insert_old_data():
    #used when a user logs in so that the data is automatically loaded in
    insertOldData()
    downloadExistingData()
    return redirect("/selections")

def insertOldData(oldData = None):
    if oldData == None:
        try:
            with open("deliverables/oldData.txt", "r") as file:
                oldData = file.read()
        except FileNotFoundError:
            return
   
    global tutors
    global all_courses

    tutors = {}
    all_courses = {}


    while(len(oldData) > 0):                                                #while there is still data to be parsed...
        name = oldData[:oldData.index(';')]                                 #grab name (used to be identified with vnumber)
        oldData = oldData[ oldData.index(';') + 1:]
        
        name = oldData[:oldData.index(';')]                                 #grab name
        oldData = oldData[ oldData.index(';') + 1:]

        courseIDs = ast.literal_eval(oldData[:oldData.index(';')])          #grab courses
        oldData = oldData[ oldData.index(';') + 1:]

        availabilities = ast.literal_eval(oldData[:oldData.index(';')])     #grab availabilities
        oldData = oldData[ oldData.index(';') + 1:]

        tutors[name] = Tutor(name, courseIDs, availabilities)               #create tutor with the above credentials

    update_all_courses()    #create the all_courses data structure

def updateParsedData():
    global parsedData
    parsedData = {}

    for tutor in tutors.values():

        for course in tutor.courses:
            courseSubject = course[:4]
            courseNumber = course[4:]

            if courseSubject not in parsedData:
                parsedData[courseSubject] = {}
            
            if courseNumber not in parsedData[courseSubject]:
                parsedData[courseSubject][courseNumber] = []

            parsedData[courseSubject][courseNumber].append((tutor.name, tutor.availabilities))

def downloadExistingData(fileName = "oldData.txt"):
    #simply creates the oldData string, and stores it into oldData.txt
    oldData = ""
    for tutor in tutors.items():
        oldData += f"{tutor[0]};{tutor[1].name};{tutor[1].courses};{tutor[1].availabilities};"
    try:
        # Try to open the file in exclusive creation mode
        with open(f"deliverables/{fileName}", "x") as file:
            # Write the contents of oldData to the file
            file.write(oldData)
    except FileExistsError:
        # The file already exists, open it in write mode and overwrite its contents
        with open(f"deliverables/{fileName}", "w") as file:
            file.write(oldData)

def downloadDoc():

    document = docx.Document()
    for subject in all_courses:
        # Add a header with the subject name
        document.add_heading(subject, level=1)

        course_sets = []
        #create table header as 'subject'
        for number in all_courses[subject]:
            course_sets.append({'course_id':number, 'time':combineDays(all_courses[subject][number])})


        # Add a table with headers
        table = document.add_table(rows=1, cols=2)
        hdr_cells = table.rows[0].cells
        hdr_cells[0].text = 'Course Number'
        hdr_cells[1].text = 'Days and Times Offered'

        # Set the table style to "Table Grid" to add lines
        table.style = 'Table Grid'

        # Add rows to the table for each course set
        for course_set in course_sets:
            row_cells = table.add_row().cells
            row_cells[0].text = course_set['course_id']
            row_cells[1].text = course_set['time']
    
    # Save the document
    document.save('deliverables/course_schedule.docx')

def downloadExcel():
    subject_color_mapping = {"CHEM":"807c24f0",
                            "MATH":"80ffff00","STAT":"80ffff00",
                            "BIOL":"80ff0000","PHIS":"80ff0000","HPEX":"80ff0000",
                            "CMSC":"80ffa500","EGRB":"80ffa500","CLSE":"80ffa500","EGRE":"80ffa500","EGMN":"80ffa500",
                            "PHYS":"8000FFFF",
                            "SPAN": "80ff69b4","FREN": "80ff69b4","GRMN": "80ff69b4","ITAL": "80ff69b4",
                            "ACCT":"800345fc","ECON":"800345fc","FIRE":"800345fc","SCMA":"800345fc","MKTG":"800345fc","MGMT":"800345fc","BUSN":"800345fc",
                            "POLI": "8000ff00","PSYC": "8000ff00","SOCY": "8000ff00","HIST": "8000ff00","PHIL": "8000ff00","ARTH": "8000ff00"}

    students = {"sunday":{}, "monday":{}, "tuesday":{}, "wednesday":{}, "thursday":{}, "friday":{}, "saturday":{}}

    #A little hard to follow, but it splits each day of the week up, and grabs all the times of each course that matches with the 'day of the week'. then append the name, for the cell
    for dayOfTheWeek in students:
        for subject in parsedData:
            color = generate_random_color()
            for number in parsedData[subject]:
                className = subject + str(number)
                students[dayOfTheWeek][className] = {}
                
                for name, dayTimes in parsedData[subject][number]:
                    for dayTime in dayTimes:
                        day = dayTime[0]
                        time = dayTime[1]
                        if day == dayOfTheWeek:
                            time_range = f"{time} - {int(time) + 1}" 
                            if time_range in students[dayOfTheWeek][className]:
                                students[dayOfTheWeek][className][time_range].append(name)
                            else:
                                students[dayOfTheWeek][className][time_range] = [name]

    # Create a new workbook
    workbook = Workbook()

    # Remove the default "Sheet" tab
    default_sheet = workbook["Sheet"]
    workbook.remove(default_sheet)

    print(f"STUDENT.ITEMS - {students.items()}")
    for day, classes in students.items():
        # Create a new sheet for each day
        sheet = workbook.create_sheet(title=day)

        # Extract class names and times from each day's classes
        class_names = list(classes.keys())
        class_names = sort_courses(class_names)
        times = []
        for class_data in classes.values():
            times.extend(class_data.keys())

        times = list(set(times))  # Remove duplicates and sort the times

        # Sort the times in ascending order
        times = sorted(times, key=lambda x: int(x.split('-')[0]))

        # Set column headers with times
        headers = ['']
        for time in times:
            start_time, end_time = time.split('-')
            headers.append(f'{start_time.strip()} - {end_time.strip()}')
        sheet.append(headers)

        # Populate rows with class names, colors, and student names
        for class_name in class_names:
            class_data = classes[class_name]
            subject = class_name[:4]

            # Set class name as row header and assign color
            if subject in subject_color_mapping:
                color_hex = subject_color_mapping[subject]  # Default to white if color is not found
            else:
                color_hex = 'ffffffff'    #if subject is not found, make it white
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            sheet.cell(row=len(sheet['A']) + 1, column=1).fill = fill
            sheet.cell(row=len(sheet['A']), column=1, value=class_name)

            # Set student names for each time
            for i, time in enumerate(times, start=2):
                start_time, end_time = time.split('-')
                time_range = f'{start_time.strip()} - {end_time.strip()}'
                students_list = class_data.get(time_range, [])
                cell = sheet.cell(row=len(sheet['A']), column=i)
                cell.value = ", ".join(students_list)

        # Adjust cell widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2  # Adding extra width for padding
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Adjust cell heights
        for row in sheet.rows:
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    adjusted_height = lines * 14  # Assuming 14 points for each line
                    if adjusted_height > max_height:
                        max_height = adjusted_height
            sheet.row_dimensions[row[0].row].height = max_height

    # Save the workbook
    workbook.save('deliverables/schedule.xlsx')


    return "done"

def combineDays(timesSet):
    #takes combined days and times ((monday, 1),(tuesday, 2)) and makes them readable - (monday 1-2pm, tuesday 2-3pm)
    timesDict = {}
    result = ""
    for time in timesSet:
        day = time[0]
        hour = time[1]

        if day not in timesDict:
            timesDict[day] = {hour}
        else:
            timesDict[day].add(hour)
        
    for day in timesDict:
        result += f"{day}: {combineTimes(timesDict[day])}\n"

    return result

def combineTimes(times):
    #used to combine times into something readable
    #so (1,2,3) turns into 1pm-4pm
    tempTimes = set({})   #all times represent the start time, so we also need to add the end times ... (1,2) should be (1,2,3)
    for time in times:    #we use the tempTimes because we cannot chnage the size of the set while interating through it
        time = int(time)
        tempTimes.add(time)

    times = list(sorted(tempTimes)) # sort the list of numbers in ascending order
    ranges = []  # create an empty list to store the condensed ranges
    start = end = times[0]  # initialize the start and end of the first range
    for time in times[1:]:
        if time == end + 1:  # if the current number is part of the current range
            end = time  # update the end of the current range
        else:  # if the current number is not part of the current range
            start_time = convertToRegTime(start)
            end_time = convertToRegTime(end + 1)
            ranges.append((start_time, end_time))  # add the current range to the list of ranges
            start = end = time  # start a new range with the current number
    start_time = convertToRegTime(start) # add the last range to the list of ranges
    end_time = convertToRegTime(end + 1)
    ranges.append((start_time, end_time)) 
    return ',\n\t  '.join(f'{r[0]}-{r[1]}' if r[0] != r[1] else str(r[0]) for r in ranges)

#converts single digit miliraty time to regular 12 hour time
def convertToRegTime(military_time):
    dt_obj = datetime.datetime.strptime(str(military_time), "%H")
    regular_time = dt_obj.strftime("%I:%M %p").lstrip('0')
    return regular_time

def generate_random_color():
    #can be used if you want to randomly generate a color, isnt being used right now
    r = random.randint(0, 255)  # Random value for red component
    g = random.randint(0, 255)  # Random value for green component
    b = random.randint(0, 255)  # Random value for blue component
    a = random.randint(0, 255)  # Random value for alpha channel (transparency)

    # Convert RGBA values to aRGB hex format
    color_hex = '{:02x}{:02x}{:02x}{:02x}'.format(a, r, g, b)

    return color_hex

def day_to_num(day):
    #defines the order of the days in the week. Returns where each day is in respect to its position of the week 
    days = ["sunday", "monday", "tuesday", "wednesday", "thursday", "friday", "saturday"]
    return days.index(day.lower())

def sort_availabilities(availabilities):
    #sort by day then by time using the day_to_num method
    availabilities = sorted(availabilities, key=lambda x: (day_to_num(x[0]), x[1]))
    #join together to make a string
    return availabilities

def sort_courses(courses):
    #similar to sorting based on day of the week, this sorts based on this specified order of classes
    #this way, things of the same colors are put next to eachother, if a subject isnt in this list, it is put at the bottom

    course_order = ["CHEM", "MATH", "STAT", "BIOL", "PHYS", "HPEX", "CMSC", "EGRB", "CLSE", "EGMN",
                        "EGRE", "PHYS", "SPAN", "FREN", "GRMN", "ITAL", "ACCT", "ECON", "FIRE", "SCMA",
                        "MKTG", "MGMT", "BUSN", "POLI", "PSYC", "SOCY", "HIST", "PHIL", "ARTH"]

    max_index = len(course_order) - 1
    return sorted(courses, key=lambda x: course_order.index(x[:4]) if x[:4] in course_order else max_index+1)

#when a person adds a new availability or a new course, this must be run so that all_courses can be updated. 
#for example if Ted works BIOL101 and works on sunday 1-2, then he decides he wants to work BIOL102 as well, 
#then we must add 1-2 for BIOL102 in all_courses
#if no name is passed, we create a new all_tutors from the tutors list, mainly used when deleting a tutor
def update_all_courses(name = "all_tutors"):
    global all_courses
    if name == "all_tutors":                        #if we want to completely refresh all_courses
        all_courses = {}                            #reset all_coursees
        for tutor in tutors.values():                     #for every tutor....
            tutor_availabilities = tutor.availabilities   #set tutor_availabilites
            tutor_courses = tutor.courses       #set tutor_courses
            for course in tutor_courses:        #for every course a tutor teaches
                subject = course[:4]            #CMSC in CMSC101
                number = course[4:]             #101 in CMSC101

                if subject not in all_courses:  #if this subject doesnt exist, make it exist
                    all_courses[subject] = {}
                if number not in all_courses[subject]:  #if that course number doesnt exist, make availabilities the tutor_availabilities
                    all_courses[subject][number] = tutor_availabilities
                else:                                   #otherwise find the union of the two
                    all_courses[subject][number] = all_courses[subject][number] | tutor_availabilities

    else:           #if we only want to insert data into a single tutor, we can go through only that tutors data rather than updating everything again
        #The process here is the same as above, only we access just one tutor (the one whos data has been added to)
        tutor_courses = tutors[name].courses
        tutor_availabilities = tutors[name].availabilities

        for course in tutor_courses:
            subject = course[:4]
            number = course[4:]

            if subject not in all_courses:
                    all_courses[subject] = {}
            if number not in all_courses[subject]:
                all_courses[subject][number] = tutor_availabilities
            else:
                all_courses[subject][number] = all_courses[subject][number] | tutor_availabilities

        