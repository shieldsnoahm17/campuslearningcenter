import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def create_excel_from_xml(xml_file):
    # Load XML file
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Create a new workbook and sheet
    workbook = Workbook()
    sheet = workbook.active

    # Define colors for each class
    class_colors = {
        'BIO101': 'FF0000',
        'ENG201': '00FF00',
        'MAT301': '0000FF',
        # Add more classes and colors as needed
    }

    # Set column headers with times
    column_index = 2  # Start from column B
    for child in root[0]:
        time = child.attrib['time']
        sheet.cell(row=1, column=column_index, value=time)
        column_index += 1

    # Populate rows with class names
    row_index = 2  # Start from row 2
    for child in root:
        class_name = child.attrib['name']
        sheet.cell(row=row_index, column=1, value=class_name)

        # Set color for the class
        if class_name in class_colors:
            color = class_colors[class_name]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sheet.cell(row=row_index, column=1).fill = fill

        # Set class names for each time
        column_index = 2
        for time_child in child:
            student_name = time_child.attrib['name']
            sheet.cell(row=row_index, column=column_index, value=student_name)
            column_index += 1

        row_index += 1

    # Save the workbook
    workbook.save('classes.xlsx')

# Usage
xml_file_path = 'classes.xml'
create_excel_from_xml(xml_file_path)
